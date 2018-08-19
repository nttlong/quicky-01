class dropdowndata(object):
    def __init__(self,owner):
        self.name =""
        self.owner = owner
        self.data =[]
        self.col = None
    def set_data(self,data):
        for x in data:
            self.owner.ws.append(x)
        return self

class column_config(object):
    def __init__(self,owner):
        self.owner = owner
        self.field = None
        self.caption = None
        self.is_hidden = False
        self.dropdown = None
        self.is_ref_list =False
        self._ref_dropdown = None
        self.data_type = "text"
    def init_data(self,*args,**kwargs):
        if type(args) is tuple:
            self.field = args[0]
            self.caption = args[1]
            self.depth_fields=self.field.split('.')
            self.depth_count=self.depth_fields.__len__()
            if args.__len__()>2:
                self.is_hidden =args[2]
        return self
    def dropdown_list(self,caption = None,data=[]):
        if self.dropdown == None:
            self.dropdown = dropdowndata(self)
            self.dropdown.name = "ref_"+self.field
            self.ws = self.owner.wb.create_sheet("ref_"+self.field)
            self.ws.append(data)
            index_col = self.owner.columns.index(self)
            _col = column_config(self.owner)
            _col.is_ref_list= True
            _col.caption =caption
            _col.col = self


            self.owner.columns.insert(index_col+1,_col)
    def set_data_type(self,type_name):
        if not type_name in ["text","number","bool","date"]:
            raise (Exception("type name must be in :\n"
                             "\t\ttext\n"
                             "\t\tnumber\n"
                             "\t\tbool\n"
                             "\t\tdate\n"))
        self.data_type = type_name
class config(object):
    def __init__(self):
        self.columns=[]
        self.wb = None
        self.ws = None
        self.data =[]
        self.has_create_name_range = False
        from openpyxl import Workbook
        from openpyxl import utils
        if self.wb == None:
            self.wb = Workbook()
        fx = [x for x in self.wb.worksheets if x.title == "data"]
        if fx.__len__() == 0:
            self.ws = self.wb.active
            self.ws.title = "data"
    def init_data(self,*args,**kwargs):
        from openpyxl import Workbook
        from openpyxl import utils
        for x in args:
            col = column_config(self)
            self.columns.append(
                col.init_data(*x)
            )


        return self
    @property
    def workbook(self):
        return self.wb
    def create_name_range(self):
        if self.has_create_name_range:
            return

        from openpyxl import Workbook
        from openpyxl import utils
        caption_list = [x.caption for x in self.columns]
        field_list = [x.field for x in self.columns]
        self.ws.append(caption_list)
        col_index = 1
        for x in self.columns:
            if x.field != None:
                self.wb.create_named_range(
                    x.field,
                    self.ws,
                    "$" + utils.get_column_letter(col_index) + ":$" +
                    utils.get_column_letter(col_index)
                )
            col_index = col_index + 1
        self.has_create_name_range = True
        return self
    def create_dropdown_cols(self):
        from openpyxl import utils
        from openpyxl.worksheet.datavalidation import DataValidation
        idx =1
        for x in self.columns:
            if x.dropdown != None:

                dv = DataValidation(type="list", formula1=x.dropdown.name+'!$A:$A', allow_blank=True)
                data_len=list(self.ws.rows).__len__()
                address_col = utils.get_column_letter(idx)+"2:"+utils.get_column_letter(idx)+"1048576"
                print address_col
                dv.add(address_col)
                self.ws.add_data_validation(dv)
            idx = idx + 1
        return self
    def create_v_lookup(self):
        rows = list(self.ws.rows)
        idx =0
        for x in self.columns:
            if x.is_ref_list:
                from openpyxl.formula import Tokenizer
                fx = r"=VLOOKUP(A{0},'{1}'!$A:$B,2)"
                for i in range(1,rows.__len__()):
                    rows[i][idx].value = fx.format(i+1,x.col.dropdown.name).upper()
            idx = idx +1
        return self


    def save(self,path_to_file):
        self.create_name_range()
        self.fetch_data(self.data)
        self.create_dropdown_cols()
        self.create_v_lookup()
        self.wb.save(path_to_file)
        return self
    def exract_item(self,item):
        ret_list =[]
        for col in self.columns:
            if col.is_ref_list:
                ret_list.append(None)
            else:
                ret = item
                for i in range(0,col.depth_count-1):
                    ret=ret.get(col.depth_fields[i],{})
                    if ret == None:
                        ret = {}
                if type(ret) is list:
                    val = "[..]"
                    ret_list.append(val)
                else:
                    val = ret.get(col.depth_fields[col.depth_count-1],None)
                    if val != None and hasattr(val,"_ObjectId__id"):
                        val = val.__str__()
                    if not type(val) is dict:
                        if type(val) is list:
                            val ="[..]"
                            ret_list.append(val)
                        else:
                            ret_list.append(val)
        return ret_list
    def unwind_data(self,items):
        for x in items:
            yield self.exract_item(x)
    def set_data(self,data):
        self.data =data
        return self
    def fetch_data(self,items):
        if items.__len__() == 0:
            items.append({})
        from openpyxl import Workbook
        from openpyxl import utils
        if self.wb == None:
            self.wb = Workbook()
        fx = [x for x in self.wb.worksheets if x.title == "data"]
        if fx.__len__() == 0:
            self.ws = self.wb.create_sheet("data",1)
        self.ws.title = "data"


        for row in self.unwind_data(items):
            self.ws.append(row)
        return self
    def cols(self,name,caption = None):
        cols = [x for x in self.columns if x.field == name]
        if cols.__len__() ==0:
            col = column_config(self)
            col.init_data(*(name,caption))
            self.columns.append(col)
            from openpyxl import Workbook
            from openpyxl import utils
            col_count =list(self.ws.columns).__len__()
            self.ws.insert_cols(col_count,1)
            return col
        else:
            return cols[0]


def create(*args,**kwargs):
    ret = config()
    ret.init_data(*args,**kwargs)
    return ret




